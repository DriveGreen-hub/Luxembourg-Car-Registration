#!/usr/bin/env python3
"""
build_data.py — Luxembourg new-vehicle registration pipeline
============================================================
Turns the monthly "Parc Automobile du Luxembourg" fleet snapshot (a ~170 MB
XLSX / ~840 MB XML export of *every* registered vehicle) into the small,
integer-coded JSON the dashboard reads (data/registrations.json).

Registration flows are reconstructed from a stock snapshot using each
vehicle's dates:
  DATCIR_GD  first registration in Luxembourg  -> the month bucket
  DATCIRPRM  first registration anywhere
  brand-new  -> DATCIRPRM == DATCIR_GD   (first-ever registration was in LU)
  import     -> DATCIRPRM <  DATCIR_GD   (registered abroad first)

USAGE
  pip install requests openpyxl
  python build_data.py                 # latest snapshot, full reconstruction
  python build_data.py --months 72     # keep only last 72 months
  python build_data.py --file p.xlsx   # use an already-downloaded file
  python build_data.py --append        # merge fresh month into existing JSON
"""
import argparse, json, os, re, sys, datetime as dt
from collections import defaultdict

DATASET_API = "https://data.public.lu/api/1/datasets/59cbac9f111e9b6be027c292/"
CACHE = "cache"
OUT = "data/registrations.json"
DRIVETRAINS = ["BEV", "PHEV", "HEV", "Petrol", "Diesel", "Other"]

# Columns we read (canonical SNCA codes). Loose matching handles case/spacing.
NEED = ["CATSTC", "CATEU", "LIBCAR", "CODEOP", "LIBMRQ", "TYPCOM", "LIBCRB", "CODCRB",
        "DATCIRPRM", "DATCIR_GD", "AUTOELEC", "CONSELEC",
        "CO2WLTP", "INFCO2"]
ESSENTIAL = ["CATSTC", "DATCIR_GD", "LIBMRQ"]   # without these we cannot proceed

# national category (CATSTC) -> dashboard segment
SEGMENT_BY_CAT = {}
for c in [1]:                          SEGMENT_BY_CAT[c] = "car"
for c in [32, 33]:                     SEGMENT_BY_CAT[c] = "van"
for c in [71, 72, 73, 74, 75, 76, 89]: SEGMENT_BY_CAT[c] = "bus"

SEGMENTS = ["car", "van", "bus", "hdv"]   # display / storage order

# Bump whenever parsing/classification logic changes in a way that should force a
# full rebuild of already-processed months (the backfill compares this to the value
# stored in the data file and starts clean on a mismatch).
DATA_VERSION = 3

def segment_of(catstc, cateu):
    """car/van/bus come from the validated national code (CATSTC);
    HDV (heavy goods, >3.5t) comes from the EU category N2/N3, which is the
    only field that cleanly separates heavy trucks from vans and trailers."""
    eu = str(cateu or "").strip().upper()
    if eu.startswith("N2") or eu.startswith("N3"):
        return "hdv"
    try:
        return SEGMENT_BY_CAT.get(int(_num(catstc)))
    except Exception:
        return None


def _norm(s):
    """Normalise a header for loose matching: upper-case, alnum only."""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def classify_drivetrain(libcrb, codcrb, autoelec, conselec):
    """Classify drivetrain. Primary signal is the unambiguous fuel CODE (CODCRB);
    the human label (LIBCRB) is only a fallback.
       PEV       -> BEV   (pure electric)
       OVC-HEV   -> PHEV  (Off-Vehicle Charging = plug-in hybrid)
       NOVC-HEV  -> HEV   (Not OVC = regular/mild hybrid)
       STD       -> Petrol/Diesel
       DUAL      -> Other (bi-fuel, e.g. LPG/CNG)
    """
    code = str(codcrb if codcrb is not None else "").upper().replace(" ", "")
    s = str(libcrb if libcrb is not None else "").upper()
    if code.startswith("PEV") or code in ("BEV", "EV"):
        return "BEV"
    if code.startswith("NOVC-HEV") or code.startswith("NOVCHEV"):
        return "HEV"
    if code.startswith("OVC-HEV") or code.startswith("OVCHEV"):
        return "PHEV"
    if code.startswith("STD"):
        return "Diesel" if "DIESEL" in s else "Petrol"
    if code.startswith("DUAL"):
        return "Other"
    # ---- fallback: parse the label, hybrids BEFORE pure-electric so that
    #      'Hybride Electrique …' is never mistaken for a BEV ----
    if "PLUG-IN" in s or "RECHARGEABLE" in s:
        return "PHEV"
    if "HYBRID" in s:
        return "PHEV" if _num(autoelec) > 0 else "HEV"
    if "PUR ELECTR" in s or s.strip() in ("ELECTRIQUE", "ELECTRIC", "ELECTRICITE"):
        return "BEV"
    if "ELECTR" in s and ("ESSENCE" in s or "DIESEL" in s):  # combined => hybrid
        return "HEV"
    if "ELECTR" in s:
        return "BEV"
    if "DIESEL" in s or "GASOIL" in s or "GAZOLE" in s:
        return "Diesel"
    if "ESSENCE" in s or "PETROL" in s or "BENZIN" in s:
        return "Petrol"
    if "LPG" in s or "GPL" in s or "CNG" in s or "GAZ" in s:
        return "Other"
    return "Other"


def classify_operation(codeop, d_first, d_lu):
    """new vs import, using the authoritative operation code when available.
       N  -> new registration
       I  -> import (used vehicle previously registered abroad)
       E/E1/H -> export/suspension: not an in-fleet registration, skip
       T/blank/other -> resold car whose code drifted; recover the original
                        nature from the dates (needed for historical months).
    Returns 'new', 'import', or None (skip).
    """
    c = str(codeop if codeop is not None else "").strip().upper()
    if c == "N":
        return "new"
    if c == "I":
        return "import"
    if c in ("E", "E1", "H"):
        return None
    # No authoritative code — older snapshots often leave CODEOP blank or numeric.
    # Treat as an import only when there is a REAL prior-registration gap: a brand-new
    # car has first-ever registration == first-LU registration, but processing lag can
    # leave a few days between them, which must not be mistaken for a foreign history.
    return "import" if (d_lu - d_first).days > 31 else "new"


def _num(x):
    try:
        return float(str(x).replace(",", ".").strip())
    except Exception:
        return 0.0


def _parse_date(x):
    if x is None or x == "":
        return None
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    # Excel serial date: days since 1899-12-30 (Windows epoch). Older snapshots can
    # ship date columns as bare numbers when the cell lost its date formatting.
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        n = int(x)
        if 20000 <= n <= 80000:                # ~1954 .. ~2089, sane reg-date range
            try:
                return dt.date(1899, 12, 30) + dt.timedelta(days=n)
            except Exception:
                return None
        return None
    x = str(x).strip()
    if not x:
        return None
    if x.isdigit() and 4 < len(x) <= 5:        # numeric string that is really a serial
        n = int(x)
        if 20000 <= n <= 80000:
            return dt.date(1899, 12, 30) + dt.timedelta(days=n)
    for f in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%Y%m%d", "%d-%m-%Y",
              "%d/%m/%y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(x[:10], f).date()
        except ValueError:
            continue
    return None


# ============================ download ======================================
def resolve_latest_xlsx():
    import requests
    print("· querying dataset API for the latest snapshot…")
    r = requests.get(DATASET_API, timeout=60)
    r.raise_for_status()
    best = None
    for it in r.json().get("resources", []):
        title = (it.get("title") or "")
        m = re.search(r"(\d{6})", title)
        if title.lower().endswith(".xlsx") and m:
            ym = m.group(1)
            if best is None or ym > best[0]:
                best = (ym, it["url"], title)
    if not best:
        sys.exit("Could not find an XLSX resource in the dataset.")
    print(f"· latest: {best[2]}  ({best[0]})")
    return best


def list_all_xlsx():
    import requests
    r = requests.get(DATASET_API, timeout=60)
    r.raise_for_status()
    out = {}
    for it in r.json().get("resources", []):
        title = (it.get("title") or "")
        m = re.search(r"(\d{6})", title)
        if title.lower().endswith(".xlsx") and m:
            out[m.group(1)] = it["url"]
    return out


def download(url, dest):
    import requests
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    if os.path.exists(dest) and os.path.getsize(dest) > 1_000_000:
        print(f"· using cached {dest}")
        return dest
    print(f"· downloading {url}")
    with requests.get(url, stream=True, timeout=900) as r:
        r.raise_for_status()
        done = 0
        with open(dest, "wb") as f:
            for chunk in r.iter_content(1 << 20):
                f.write(chunk); done += len(chunk)
                print(f"\r  {done/1e6:6.1f} MB", end="")
    print()
    return dest


# ============================ parse =========================================
def iter_rows_xlsx(path):
    """Stream rows from the data sheet, matching headers loosely.

    The SNCA export ships a wrong/minimal <dimension> tag, which makes
    openpyxl's read-only mode read only the first column. We bypass that by
    passing an explicit wide column range (max_col) to iter_rows.
    """
    from openpyxl import load_workbook
    PROBE = 256                                   # read up to this many columns
    wb = load_workbook(path, read_only=True, data_only=True)

    def header_of(ws):
        first = next(ws.iter_rows(min_row=1, max_row=1, max_col=PROBE, values_only=True), None)
        return list(first) if first else []

    # Pick the sheet whose first row matches the most known columns.
    best = None
    for name in wb.sheetnames:
        hdr = header_of(wb[name])
        hn = [_norm("" if h is None else h) for h in hdr]
        matched = sum(1 for c in NEED if _norm(c) in hn)
        if best is None or matched > best[0]:
            best = (matched, name, hdr, hn)
    matched, name, hdr, hn = best

    essential_found = [c for c in ESSENTIAL if _norm(c) in hn]
    if len(essential_found) < len(ESSENTIAL):
        seen = [h for h in hdr if h not in (None, "")]
        wb.close()
        raise SystemExit(
            "\n*** COLUMN MISMATCH — cannot parse this file ***\n"
            f"Best sheet: '{name}' (matched {matched}/{len(NEED)} known columns).\n"
            f"Essential columns found: {essential_found} of {ESSENTIAL}.\n"
            f"HEADERS SEEN ({len(seen)}): {seen}\n"
            "Copy the HEADERS SEEN line above so the column mapping can be fixed.\n")

    ncols = max([i for i, h in enumerate(hdr) if h not in (None, "")], default=-1) + 1
    idx = {c: hn.index(_norm(c)) for c in NEED if _norm(c) in hn}
    print(f"· data sheet '{name}': matched {matched}/{len(NEED)} columns, {ncols} cols wide")

    ws = wb[name]
    # Large max_row + explicit max_col defeat the broken <dimension>; read_only
    # still stops at the real end of data, so the high bound is harmless.
    for row in ws.iter_rows(min_row=2, max_row=5_000_000, max_col=ncols, values_only=True):
        if row is None:
            continue
        yield {c: (row[idx[c]] if c in idx and idx[c] < len(row) else None) for c in NEED}
    wb.close()


def iter_rows_xml(path):
    import xml.etree.ElementTree as ET
    rec = {}
    for ev, el in ET.iterparse(path, events=("end",)):
        tag = el.tag.upper()
        if tag == "OPE":                 # XML uses <OPE>; xlsx uses CODEOP
            rec["CODEOP"] = el.text
        elif tag in NEED:
            rec[tag] = el.text
        elif el.tag.lower() == "vehicle":
            yield {c: rec.get(c) for c in NEED}
            rec = {}; el.clear()


def iter_rows(path):
    return iter_rows_xml(path) if path.lower().endswith(".xml") else iter_rows_xlsx(path)


# ============================ aggregate =====================================
def build(path, snapshot_ym, keep_months=None):
    counts = defaultdict(int); co2sum = defaultdict(float); co2n = defaultdict(int)
    n_read = n_kept = 0
    for v in iter_rows(path):
        n_read += 1
        if n_read % 100000 == 0:
            print(f"\r  parsed {n_read:,} rows…", end="")
        seg = segment_of(v.get("CATSTC"), v.get("CATEU"))
        if seg is None:
            continue
        d_lu = _parse_date(v.get("DATCIR_GD"))
        if d_lu is None:
            continue
        ym = f"{d_lu.year:04d}-{d_lu.month:02d}"
        d_first = _parse_date(v.get("DATCIRPRM")) or d_lu
        op = classify_operation(v.get("CODEOP"), d_first, d_lu)
        if op is None:
            continue
        brand = str(v.get("LIBMRQ") or "").strip().title() or "UNKNOWN"
        model = str(v.get("TYPCOM") or "").strip().upper() or "—"
        dtrain = classify_drivetrain(v.get("LIBCRB"), v.get("CODCRB"),
                                     v.get("AUTOELEC"), v.get("CONSELEC"))
        key = (ym, seg, op, brand, model, dtrain)
        counts[key] += 1
        co2 = _num(v.get("CO2WLTP")) or _num(v.get("INFCO2"))
        if dtrain == "BEV":
            co2sum[key] += 0.0; co2n[key] += 1
        elif co2 > 0:
            co2sum[key] += co2; co2n[key] += 1
        n_kept += 1
    print(f"\r  parsed {n_read:,} rows · {n_kept:,} cars/vans/buses kept")

    if n_kept == 0:
        raise SystemExit(
            "\n*** 0 vehicles matched — aborting so a bad file is NOT deployed. ***\n"
            f"Read {n_read:,} rows but none were cars/vans/buses with a valid "
            "registration date. Likely a value-format issue (category codes or "
            "date format). Check a few sample rows of the source file.\n")

    months = sorted({k[0] for k in counts})
    if keep_months:
        months = months[-keep_months:]
    mset = set(months); months_i = {m: i for i, m in enumerate(months)}
    segs = SEGMENTS; ops = ["new", "import"]
    brands, b_i = [], {}; models, m_i = [], {}; rows = []
    for (ym, seg, op, brand, model, dtrain), n in counts.items():
        if ym not in mset:
            continue
        if brand not in b_i:
            b_i[brand] = len(brands); brands.append(brand)
        mk = (b_i[brand], model)
        if mk not in m_i:
            m_i[mk] = len(models); models.append([model, b_i[brand]])
        key = (ym, seg, op, brand, model, dtrain)
        rows.append([months_i[ym], segs.index(seg), ops.index(op), b_i[brand],
                     m_i[mk], DRIVETRAINS.index(dtrain), n,
                     int(round(co2sum[key])), co2n[key]])

    return {
        "meta": {
            "generated": dt.date.today().isoformat(),
            "source": "Parc Automobile du Luxembourg (SNCA) via data.public.lu — CC0",
            "source_snapshot": snapshot_ym,
            "data_version": DATA_VERSION,
            "latest_month": months[-1] if months else None,
            "note": "Brand-new = first LU registration equals first-ever registration. "
                    "Import = vehicle previously registered abroad. Older months are "
                    "subject to survivorship bias (reconstructed from one stock snapshot).",
        },
        "dims": {"months": months, "segments": segs, "operations": ops,
                 "drivetrains": DRIVETRAINS, "brands": brands, "models": models},
        "rows": rows,
    }


def write(obj):
    os.makedirs("data", exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(obj, f, separators=(",", ":"), ensure_ascii=False)
    print(f"· wrote {OUT}  ({os.path.getsize(OUT)/1024:.0f} KB · "
          f"{len(obj['rows'])} rows · {len(obj['dims']['months'])} months · "
          f"latest {obj['meta']['latest_month']})")


# ============================ append/merge ==================================
def merge_append(old, new):
    old_src = (old.get("meta", {}).get("source", "") or "").lower()
    if "synthetic" in old_src or "sample" in old_src:
        print("· existing data is the sample — replacing entirely (no merge).")
        return new
    def explode(o):
        d = o["dims"]; out = defaultdict(lambda: [0, 0, 0])
        for r in o["rows"]:
            key = (d["months"][r[0]], d["segments"][r[1]], d["operations"][r[2]],
                   d["brands"][r[3]], d["models"][r[4]][0], d["drivetrains"][r[5]])
            out[key][0] += r[6]
            out[key][1] += r[7] if len(r) > 7 else 0
            out[key][2] += r[8] if len(r) > 8 else 0
        return out
    om, nm = explode(old), explode(new)
    new_months = {k[0] for k in nm}
    merged = {k: v for k, v in om.items() if k[0] not in new_months}
    merged.update(nm)
    months = sorted({k[0] for k in merged})
    segs = SEGMENTS; ops = ["new", "import"]
    mi = {m: i for i, m in enumerate(months)}
    brands, bi = [], {}; models, midx = [], {}; rows = []
    for (ym, seg, op, brand, model, dtr), vals in merged.items():
        if brand not in bi:
            bi[brand] = len(brands); brands.append(brand)
        mk = (bi[brand], model)
        if mk not in midx:
            midx[mk] = len(models); models.append([model, bi[brand]])
        rows.append([mi[ym], segs.index(seg), ops.index(op), bi[brand],
                     midx[mk], DRIVETRAINS.index(dtr), vals[0], vals[1], vals[2]])
    new["dims"] = {"months": months, "segments": segs, "operations": ops,
                   "drivetrains": DRIVETRAINS, "brands": brands, "models": models}
    new["rows"] = rows
    new["meta"]["latest_month"] = months[-1]
    return new


# ============================ main ==========================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file")
    ap.add_argument("--months", type=int, default=None)
    ap.add_argument("--append", action="store_true")
    ap.add_argument("--categories", action="store_true",
                    help="just print the CATSTC categories + labels + counts and exit")
    ap.add_argument("--inspect", metavar="YYYYMM", default=None,
                    help="download ONE snapshot and dump its structure + sample rows, then exit")
    args = ap.parse_args()

    if args.inspect:
        snaps = list_all_xlsx()
        keys = sorted(snaps)
        print(f"· {len(keys)} snapshots available: {keys[0]} … {keys[-1]}")
        ym = args.inspect
        if ym not in snaps:
            later = [k for k in keys if k >= ym]
            ym = later[0] if later else keys[-1]
            print(f"· {args.inspect} not found; using nearest = {ym}")
        path = download(snaps[ym], os.path.join(CACHE, f"inspect-{ym}.xlsx"))
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        print(f"\n=== snapshot {ym} ===\nsheets: {wb.sheetnames}")
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(min_row=1, max_row=4, max_col=256, values_only=True))
            print(f"\n--- sheet '{name}' (showing header + 3 rows) ---")
            if rows:
                hdr = [(i, h) for i, h in enumerate(rows[0]) if h not in (None, "")]
                print(f"  header ({len(hdr)} non-empty): {hdr[:45]}")
            for r in rows[1:4]:
                cells = [(i, repr(v), type(v).__name__)
                         for i, v in enumerate(r) if v not in (None, "")][:45]
                print(f"  row: {cells}")
        print("\n-- parsed key fields via iter_rows (first 5 rows; shows value + python type) --")
        try:
            for i, v in enumerate(iter_rows(path)):
                if i >= 5:
                    break
                print({k: (repr(v.get(k)), type(v.get(k)).__name__)
                       for k in ["CATSTC", "CATEU", "DATCIR_GD", "DATCIRPRM",
                                 "CODEOP", "LIBMRQ", "LIBCRB"]})
        except SystemExit as e:
            print(f"  iter_rows could not parse this file: {e}")
        return

    if args.file:
        path = args.file
        m = re.search(r"(\d{6})", os.path.basename(path))
        snapshot = m.group(1) if m else "local"
    else:
        ym, url, _ = resolve_latest_xlsx()
        path = download(url, os.path.join(CACHE, f"parc-{ym}.xlsx"))
        snapshot = ym

    if args.categories:
        from collections import Counter
        cateu = Counter(); catstc = Counter(); label = {}; seg_tot = Counter()
        for v in iter_rows(path):
            eu = str(v.get("CATEU") or "").strip().upper()
            cs = str(v.get("CATSTC"))
            cateu[eu] += 1
            catstc[cs] += 1
            if cs not in label:
                label[cs] = str(v.get("LIBCAR") or "")
            seg_tot[segment_of(v.get("CATSTC"), v.get("CATEU")) or "(skipped)"] += 1

        print("\n=== CATEU — EU vehicle category (drives HDV) ===")
        print("  M1=car  N1=van  N2/N3=heavy goods(HDV)  M2/M3=bus  L=moto  O=trailer  T=tractor")
        print("  " + "-" * 58)
        for k, n in cateu.most_common(50):
            hdv = " <-- HDV" if (k.startswith("N2") or k.startswith("N3")) else ""
            print(f"  {n:9d}  {(k or '(blank)'):8}{hdv}")

        print("\n=== CATSTC — national code + bodywork label ===")
        print("  " + "-" * 58)
        for k, n in catstc.most_common(60):
            try:
                seg = SEGMENT_BY_CAT.get(int(float(k)))
            except Exception:
                seg = None
            print(f"  {n:9d}  CATSTC={k:>4}  {label[k][:34]:34}  {seg or '(via CATEU / skipped)'}")

        print("\n=== resulting dashboard segments (what actually gets kept) ===")
        print("  " + "-" * 58)
        for k in ["car", "van", "bus", "hdv", "(skipped)"]:
            if seg_tot.get(k):
                print(f"  {seg_tot[k]:9d}  {k}")
        return

    obj = build(path, snapshot, keep_months=args.months)

    if args.append and os.path.exists(OUT):
        try:
            obj = merge_append(json.load(open(OUT, encoding="utf-8")), obj)
        except Exception as e:
            print(f"· append skipped ({e}); writing fresh data.")

    write(obj)


if __name__ == "__main__":
    main()
